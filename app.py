"""
Piping Business Agent â Web Dashboard
Run locally:    streamlit run app.py
Cloud:          Deploy to Streamlit Community Cloud (free)
"""
import streamlit as st
import pandas as pd
import tempfile
import io
from pathlib import Path
from datetime import date, datetime, timedelta
import json, sys, os

ROOT = Path(__file__).resolve().parent
os.chdir(ROOT)
sys.path.insert(0, str(ROOT / "scripts"))

# âââââââââââââââââââââ INLINE HELPERS (from scripts/common.py) âââââââââââââââââââââ
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_config():
    cfg_path = ROOT / "config.json"
    if not cfg_path.exists():
        cfg_path = ROOT / "config.example.json"
    with open(cfg_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _rows_from_sheet(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def load_inventory():
    if not HAS_OPENPYXL:
        return _demo_inventory()
    cfg = load_config()
    inv_path = ROOT / cfg["paths"]["inventory"]
    if not inv_path.exists():
        return _demo_inventory()
    wb = load_workbook(inv_path, data_only=True)
    return _rows_from_sheet(wb["Inventory"])


def load_clients():
    if not HAS_OPENPYXL:
        return _demo_clients()
    cfg = load_config()
    cl_path = ROOT / cfg["paths"]["clients"]
    if not cl_path.exists():
        return _demo_clients()
    wb = load_workbook(cl_path, data_only=True)
    return _rows_from_sheet(wb["Clients"])


def find_product(code):
    code = code.strip().upper()
    for p in load_inventory():
        if str(p.get("Product Code", "")).strip().upper() == code:
            return p
    return None


def find_client(identifier):
    identifier = identifier.strip()
    for c in load_clients():
        if (str(c.get("Client ID", "")).strip().lower() == identifier.lower() or
                str(c.get("Client Name", "")).strip().lower() == identifier.lower()):
            return c
    return None


def _demo_inventory():
    return [
        {"Product Code": "SS304-PIPE-1", "Product Name": "SS 304 Pipe 1 inch", "Category": "Pipes",
         "Stock Qty": 150, "Reorder Level": 50, "Cost Price (INR)": 320, "Selling Price (INR)": 480,
         "Supplier": "Steel India Corp", "HSN Code": "7306"},
        {"Product Code": "SS316-PIPE-2", "Product Name": "SS 316 Pipe 2 inch", "Category": "Pipes",
         "Stock Qty": 30, "Reorder Level": 40, "Cost Price (INR)": 580, "Selling Price (INR)": 870,
         "Supplier": "Metro Tubes", "HSN Code": "7306"},
        {"Product Code": "GI-ELBOW-1", "Product Name": "GI Elbow 1 inch", "Category": "Fittings",
         "Stock Qty": 200, "Reorder Level": 100, "Cost Price (INR)": 45, "Selling Price (INR)": 75,
         "Supplier": "Patel Fittings", "HSN Code": "7307"},
        {"Product Code": "SS304-TEE-1.5", "Product Name": "SS 304 Tee 1.5 inch", "Category": "Fittings",
         "Stock Qty": 80, "Reorder Level": 30, "Cost Price (INR)": 120, "Selling Price (INR)": 190,
         "Supplier": "Steel India Corp", "HSN Code": "7307"},
        {"Product Code": "CPVC-PIPE-0.5", "Product Name": "CPVC Pipe 0.5 inch", "Category": "Pipes",
         "Stock Qty": 10, "Reorder Level": 50, "Cost Price (INR)": 85, "Selling Price (INR)": 130,
         "Supplier": "Astral Ltd", "HSN Code": "3917"},
        {"Product Code": "VALVE-BALL-1", "Product Name": "Ball Valve 1 inch SS", "Category": "Valves",
         "Stock Qty": 60, "Reorder Level": 20, "Cost Price (INR)": 250, "Selling Price (INR)": 400,
         "Supplier": "Patel Fittings", "HSN Code": "8481"},
    ]


def _demo_clients():
    return [
        {"Client ID": "CL001", "Client Name": "Sharma Builders", "Company": "Sharma Constructions Pvt Ltd",
         "City": "Ahmedabad", "Phone (with country code)": "+919812345678",
         "Outstanding (INR)": 45000, "Payment Status": "Overdue",
         "Last Invoice Due Date": "2026-03-15", "Last Proposal Sent": "2026-04-01"},
        {"Client ID": "CL002", "Client Name": "Patel Industries", "Company": "Patel Industrial Solutions",
         "City": "Surat", "Phone (with country code)": "+919876543210",
         "Outstanding (INR)": 0, "Payment Status": "Paid",
         "Last Invoice Due Date": "2026-02-28", "Last Proposal Sent": "2026-03-20"},
        {"Client ID": "CL003", "Client Name": "Mehta Plumbing", "Company": "Mehta & Sons Plumbing",
         "City": "Rajkot", "Phone (with country code)": "+919898989898",
         "Outstanding (INR)": 12500, "Payment Status": "Due",
         "Last Invoice Due Date": "2026-04-10", "Last Proposal Sent": None},
        {"Client ID": "CL004", "Client Name": "Gujarat Pharma", "Company": "Gujarat Pharma Works",
         "City": "Vadodara", "Phone (with country code)": "+919123456789",
         "Outstanding (INR)": 78000, "Payment Status": "Overdue",
         "Last Invoice Due Date": "2026-03-01", "Last Proposal Sent": "2026-04-05"},
    ]

# âââââââââââââââââââââ CONFIG âââââââââââââââââââââ
cfg = load_config()
BIZ = cfg["business"]

st.set_page_config(
    page_title=f"{BIZ['name']} â Agent",
    page_icon="ð§",
    layout="wide",
    initial_sidebar_state="expanded",
)

# âââââââââââââââââââââ SIDEBAR âââââââââââââââââââââ
st.sidebar.title("ð§ " + BIZ["name"])
st.sidebar.caption(BIZ["tagline"])
page = st.sidebar.radio(
    "Navigate",
    ["ð Dashboard", "ð Generate Proposal", "ð§¾ Bill to Tally",
     "ð Delivery Challan", "ð² WhatsApp Reminders",
     "ð¦ Inventory", "ð¥ Clients"],
)

# âââââââââââââââââââââ HELPERS âââââââââââââââââââââ
def _parse_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def client_options():
    return {f"{c['Client ID']} â {c['Client Name']}": c for c in load_clients()}


def product_options():
    return {f"{p['Product Code']} â {p['Product Name']}": p for p in load_inventory()}


def product_rows_ui(key_prefix, session_key):
    """Reusable product row selector. Returns list of (product_dict, qty)."""
    popts = product_options()
    pkeys = list(popts.keys())
    if session_key not in st.session_state:
        st.session_state[session_key] = [{"product": pkeys[0], "qty": 1}]

    rows_to_delete = []
    for i, row in enumerate(st.session_state[session_key]):
        c1, c2, c3 = st.columns([5, 2, 1])
        cur_idx = pkeys.index(row["product"]) if row["product"] in pkeys else 0
        row["product"] = c1.selectbox("Product", pkeys, index=cur_idx, key=f"{key_prefix}_prod_{i}")
        row["qty"] = c2.number_input("Qty", min_value=1, value=row["qty"], key=f"{key_prefix}_qty_{i}")
        if c3.button("â", key=f"{key_prefix}_del_{i}"):
            rows_to_delete.append(i)

    for idx in sorted(rows_to_delete, reverse=True):
        st.session_state[session_key].pop(idx)
    if rows_to_delete:
        st.rerun()

    if st.button("â Add another product", key=f"{key_prefix}_add"):
        st.session_state[session_key].append({"product": pkeys[0], "qty": 1})
        st.rerun()

    return [(popts[r["product"]], r["qty"]) for r in st.session_state[session_key]]


def show_total_preview(items):
    subtotal = sum(float(p.get("Selling Price (INR)") or 0) * q for p, q in items)
    gst = subtotal * 0.18
    st.info(f"**Subtotal:** â¹{subtotal:,.0f}  |  **GST 18%:** â¹{gst:,.0f}  |  **Grand Total:** â¹{subtotal + gst:,.0f}")
    return subtotal


# âââââââââââââââââââââ PAGE: DASHBOARD âââââââââââââââââââââ
if page == "ð Dashboard":
    st.title("ð Business Dashboard")
    inv = load_inventory()
    clients = load_clients()

    total_stock_value = sum(
        float(p.get("Stock Qty") or 0) * float(p.get("Cost Price (INR)") or 0)
        for p in inv
    )
    low_stock_count = sum(
        1 for p in inv
        if float(p.get("Stock Qty") or 0) <= float(p.get("Reorder Level") or 0)
    )
    total_outstanding = sum(float(c.get("Outstanding (INR)") or 0) for c in clients)
    overdue_count = sum(
        1 for c in clients
        if str(c.get("Payment Status", "")).lower() == "overdue"
    )

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total Products", len(inv))
    k2.metric("Low Stock Items", low_stock_count,
              delta_color="inverse" if low_stock_count > 0 else "normal")
    k3.metric("Stock Value", f"â¹{total_stock_value:,.0f}")
    k4.metric("Total Outstanding", f"â¹{total_outstanding:,.0f}")

    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("â ï¸ Low Stock Alert")
        low = [p for p in inv if float(p.get("Stock Qty") or 0) <= float(p.get("Reorder Level") or 0)]
        if low:
            st.dataframe(
                pd.DataFrame(low)[["Product Code", "Product Name", "Stock Qty", "Reorder Level", "Supplier"]],
                use_container_width=True, hide_index=True,
            )
        else:
            st.success("All products above reorder level!")

    with col2:
        st.subheader("ð° Overdue / Due Payments")
        overdue = [
            c for c in clients
            if str(c.get("Payment Status", "")).lower() in ("overdue", "due")
            and float(c.get("Outstanding (INR)") or 0) > 0
        ]
        if overdue:
            st.dataframe(
                pd.DataFrame(overdue)[["Client ID", "Client Name", "Outstanding (INR)", "Payment Status", "Last Invoice Due Date"]],
                use_container_width=True, hide_index=True,
            )
        else:
            st.success("No overdue payments!")

    st.divider()
    st.subheader("ð Recent Files")
    out_dir = ROOT / "proposals_out"
    out_dir.mkdir(exist_ok=True)
    files = sorted(out_dir.glob("*"), key=lambda f: f.stat().st_mtime, reverse=True)[:10]
    if files:
        for f in files:
            col_a, col_b, col_c = st.columns([4, 2, 2])
            col_a.text(f.name)
            col_b.text(f"{f.stat().st_size / 1024:.1f} KB")
            with open(f, "rb") as fp:
                col_c.download_button("â¬ Download", fp.read(), file_name=f.name, key=f"dl_{f.name}")
    else:
        st.info("No files yet. Generate proposals or challans from the sidebar.")


# âââââââââââââââââââââ PAGE: GENERATE PROPOSAL âââââââââââââââââââââ
elif page == "ð Generate Proposal":
    st.title("ð Generate Proposal")
    st.write("Select a client, add products, click **Generate** â done.")

    copts = client_options()
    selected_client = st.selectbox("Select Client", list(copts.keys()))
    client = copts[selected_client]

    st.subheader("Products")
    items = product_rows_ui("prop", "proposal_rows")
    st.divider()
    show_total_preview(items)

    if st.button("ð Generate Proposal", type="primary"):
        with st.spinner("Generating proposal..."):
            from generate_proposal import build_proposal
            safe = "".join(ch for ch in client["Client Name"] if ch.isalnum() or ch in " _-").replace(" ", "_")
            out_path = ROOT / "proposals_out" / f"Proposal_{safe}_{date.today()}.pptx"
            out_path.parent.mkdir(exist_ok=True)
            build_proposal(client, items, cfg, out_path)

        st.success("Proposal ready!")
        with open(out_path, "rb") as fp:
            st.download_button(
                "ð¥ Download Proposal (.pptx)", fp.read(),
                file_name=out_path.name,
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            )
È8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ QÑNSÈSH8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ 8¥ [YYÙHOH¼'éï[È[HÝ]J¼'éï[È[HBÝÜ]JÜX]HHØ[\ÈÝXÚ\[ÜÝ]È[KÜÝÛØYHSÈ[\ÜX[X[KBÛÜÈHÛY[ÛÜ[ÛÊ
BÙ[XÝYØÛY[HÝÙ[XÝÞ
Ù[XÝÛY[\Ý
ÛÜËÙ^\Ê
JJBÛY[HÛÜÖÜÙ[XÝYØÛY[BÝÝXXY\[ÚXÙH][\ÈB][\ÈHÙXÝÜÝÜ×ÝZJ[H[WÜÝÜÈBÝ]Y\
BÚÝ×ÝÝ[Ü]Y]Ê][\ÊBÛÛKÛÛÛÛÈHÝÛÛ[[ÊÊBYÛÛK]Û¼'ä`{î#È]Y]ÈSNÛHÜÝÝ×Ý[H[\ÜZ[ÝÝXÚ\Þ[[Ý[HZ[ÝÝXÚ\Þ[
ÙËÛY[][\Ë]KÙ^J
JBÝ[Ê[ÚXÙHÝ[
[ÛÔÕ
N8 ®^ÝÝ[HBÝÛÙJ[[ÝXYÙOH[BYÛÛ]Û¼'äéHÝÛØYSNÛHÜÝÝ×Ý[H[\ÜZ[ÝÝXÚ\Þ[[Ý[HZ[ÝÝXÚ\Þ[
ÙËÛY[][\Ë]KÙ^J
JBÝÝÛØYØ]Û¸«!ÈØ]HS[H[[ÛÙJ]NK[WÛ[YOYÝXÚ\ÞØÛY[ÉÐÛY[Q	×_WÞÙ]KÙ^J
_K[Z[YOH\XØ][ÛÞ[
BYÛÛË]Û¼'äéÜÝÈ[H
ØØ[
H\OH[X\HNÛHÜÝÝ×Ý[H[\ÜZ[ÝÝXÚ\Þ[Ù[Ý×Ý[B[Ý[HZ[ÝÝXÚ\Þ[
ÙËÛY[][\Ë]KÙ^J
JB\HÙËÙ]
[HßJKÙ]
\ËÛØØ[ÜÝLBN\ÜHÙ[Ý×Ý[J[\
BYSQTÔ[\ÜÜVÑTSÓ[\ÜÝ\Ü[H\ÜÈÝ[ÚXÚÈYÙ\[Y\ËÜ\ÜHB[ÙNÝÝXØÙ\ÜÊÜÝYÈ[HHÝ[8 ®^ÝÝ[HB^Ù\^Ù\[Û\ÈNÝ\ÜØ[ÝXXÚ[H]Ý\K^Ü[[HÛ[Ý\ÈÚ]SÙ\\ÛÜLY[ÝIÜH\Ú[ÈHÛÝY\Ú[ÛÝÛØYHS[[\Ü]X[X[K\ÜÙ_H
B2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)HtS¢DTÄdU%4ÄÄâ)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H ¦VÆbvRÓÒ/	ù©¢FVÆfW'6ÆÆâ# ¢7BçFFÆR/	ù©¢FVÆfW'6ÆÆâ"¢7Bçw&FR$vVæW&FRDb6ÆÆâ(	B7Fö6²2WFöÖF6ÆÇFVGV7FVBâ" ¢6÷G2Ò6ÆVçEö÷Föç2¢6VÆV7FVEö6ÆVçBÒ7Bç6VÆV7F&÷%6VÆV7B6ÆVçB"ÂÆ7B6÷G2æ¶W2¢6ÆVçBÒ6÷G5·6VÆV7FVEö6ÆVçEÐ ¢7Bç7V&VFW"$FV×2FòF7F6"¢FV×2Ò&öGV7E÷&÷w5÷V&6Â"Â&6ÆÆå÷&÷w2" ¢7BæFfFW"¢7Bç7V&VFW"%G&ç7÷'BFWFÇ2"¢F3ÂF3"Ò7Bæ6öÇVÖç2"¢fV6ÆRÒF3çFWEöçWB%fV6ÆRçVÖ&W""ÂÆ6VöÆFW#Ò$t£##3B"¢G&fW"ÒF3"çFWEöçWB$G&fW"böæR"ÂÆ6VöÆFW#Ò%&×R³#3CScs"¢õ÷&VbÒ7BçFWEöçWB$6ÆVçBòò&Vb÷FöæÂ"ÂÆ6VöÆFW#Ò%òÓ##bÕ4"ÓC"" ¢b7Bæ'WGFöâ/	ù¨vVæW&FR6ÆÆâbFVGV7B7Fö6²"ÂGSÒ'&Ö'" ¢vF7Bç7ææW"$vVæW&Fær6ÆÆââââ" ¢g&öÒFVÆfW'ö6ÆÆâ×÷'B'VÆEö6ÆÆâÂöFVGV7E÷7Fö6°¢6ÆÆåöæòÒb$D2÷¶FFRçFöFç7G&gFÖRrUVÒVBrÒ÷¶6ÆVçE²t6ÆVçBBu×Ò ¢6fRÒ""æ¦öâ6f÷"6â6ÆVçE²$6ÆVçBæÖR%Òb6æ6ÆçVÒ÷"$6ÆVçB ¢÷WE÷FÒ$ôõBò'&÷÷6Ç5ö÷WB"òb$6ÆÆå÷·6fWÕ÷¶FFRçFöFÒçFb ¢÷WE÷Fç&VçBæÖ¶F"W7Eöö³ÕG'VR¢'VÆEö6ÆÆâ6frÂ6ÆVçBÂFV×2Â6ÆÆåöæòÂFFRçFöFÀ¢fV6ÆRÂG&fW"Âõ÷&VbÂ÷WE÷F¢G' ¢öFVGV7E÷7Fö6²6frÂFV×2¢7Fö6µö×6rÒ%7Fö6²FVGV7FVBg&öÒçfVçF÷'â ¢W6WBW6WFöâ2S ¢7Fö6µö×6rÒb$æ÷FS¢7Fö6²FVGV7Föâ6¶VB¶WÒ  ¢7Bç7V66W72b$6ÆÆâ¢§¶6ÆÆåöæ÷Ò¢¢vVæW&FVB·7Fö6µö×6wÒ"¢vF÷Vâ÷WE÷FÂ'&""2g ¢7BæF÷væÆöEö'WGFöâ/	ù:RF÷væÆöB6ÆÆâçFb"Âgç&VBÀ¢fÆUöæÖSÖ÷WE÷FææÖRÂÖÖSÒ&Æ6Föâ÷Fb"  ¢2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)HtS¢tE4$TÔäDU%2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H ¦VÆbvRÓÒ/	ù;"vG4&VÖæFW'2# ¢7BçFFÆR/	ù;"vG4&VÖæFW'2"¢7Bçw&FR%6VRvòv÷VÆBvWB&VÖæFW"âöæ6RGvÆò26öæfwW&VBÂ÷R6â6VæBFVÒÆfRâ" ¢FöFÒFFRçFöF¢F#ÂF#"ÂF#2Ò7BçF'2²/	ù+ÖVçG2"Â/	ù:bÆ÷r7Fö6²"Â/	ù8²&÷÷6ÂföÆÆ÷r×W%Ò ¢vFF# ¢7Bç7V&VFW"$GVRò÷fW&GVRÖVçG2"¢GVUöÆ7BÒµÐ¢f÷"2âÆöEö6ÆVçG2 ¢÷WG7FæFærÒfÆöB2ævWB$÷WG7FæFærå""÷"¢7FGW2Ò7G"2ævWB%ÖVçB7FGW2"Â""ç7G&æÆ÷vW"¢GVRÒ÷'6UöFFR2ævWB$Æ7Bçfö6RGVRFFR"¢b÷WG7FæFærÃÒ÷"7FGW2ÓÒ'B"÷"æ÷BGVS ¢6öçFçVP¢F2ÒFöFÒGVRæF0¢bF2ÂÓ# ¢6öçFçVP¢GVUöÆ7BæVæB°¢$6ÆVçB#¢5²$6ÆVçBæÖR%ÒÀ¢%öæR#¢2ævWB%öæRvF6÷VçG'6öFR"Â""À¢$÷WG7FæFær#¢b.(+¶÷WG7FæFæs¢ÂãgÒ"À¢$GVRFFR#¢7G"GVRÀ¢%7FGW2#¢$÷fW&GVR"bF2âVÇ6R$GVR6ööâ"À¢$F2#¢F2À¢Ò¢bGVUöÆ7C ¢7BæFFg&ÖRBäFFg&ÖRGVUöÆ7BÂW6Uö6öçFæW%÷vGFÕG'VRÂFUöæFWÕG'VR¢7Bææfòb'¶ÆVâGVUöÆ7BÒ6ÆVçB2v÷VÆB&V6VfRÖVçB&VÖæFW"â"¢VÇ6S ¢7Bç7V66W72$æò÷fW&GVRÖVçG2" ¢vFF## ¢7Bç7V&VFW"$Æ÷r7Fö6²FV×2"¢Æ÷rÒµÐ¢f÷"âÆöEöçfVçF÷' ¢7Fö6²ÒfÆöBævWB%7Fö6²G"÷"¢&V÷&FW"ÒfÆöBævWB%&V÷&FW"ÆWfVÂ"÷"¢b7Fö6²ÃÒ&V÷&FW# ¢Æ÷ræVæB°¢$6öFR#¢²%&öGV7B6öFR%ÒÀ¢%&öGV7B#¢²%&öGV7BæÖR%ÒÀ¢%7Fö6²ÆVgB#¢çB7Fö6²À¢%&V÷&FW"B#¢çB&V÷&FW"À¢%7WÆW"#¢ævWB%7WÆW""Â""À¢Ò¢bÆ÷s ¢7BæFFg&ÖRBäFFg&ÖRÆ÷rÂW6Uö6öçFæW%÷vGFÕG'VRÂFUöæFWÕG'VR¢7Bçv&æærb'¶ÆVâÆ÷rÒFVÒ2æVVB&V÷&FW&ærâ÷RvBvWBvG4ÆW'Bâ"¢VÇ6S ¢7Bç7V66W72$ÆÂ7Fö6²ÆWfVÇ2&RVÇF" ¢vFF#3 ¢7Bç7V&VFW"%&÷÷6Ç2vFær&W7öç6R"¢VæFærÒµÐ¢f÷"2âÆöEö6ÆVçG2 ¢6VçBÒ÷'6UöFFR2ævWB$Æ7B&÷÷6Â6VçB"¢bæ÷B6VçC ¢6öçFçVP¢F2ÒFöFÒ6VçBæF0¢b2ÃÒF2ÃÒ# ¢VæFæræVæB°¢$6ÆVçB#¢5²$6ÆVçBæÖR%ÒÀ¢%öæR#¢2ævWB%öæRvF6÷VçG'6öFR"Â""À¢%6VçBöâ#¢7G"6VçBÀ¢$F2vò#¢F2À¢Ò¢bVæFæs ¢7BæFFg&ÖRBäFFg&ÖRVæFærÂW6Uö6öçFæW%÷vGFÕG'VRÂFUöæFWÕG'VR¢7Bææfòb'¶ÆVâVæFærÒ6ÆVçB2v÷VÆBvWBföÆÆ÷r×Wâ"¢VÇ6S ¢7Bç7V66W72$æòVæFærföÆÆ÷r×W2"  ¢2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)HtS¢ådTåDõ%)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H ¦VÆbvRÓÒ/	ù:bçfVçF÷'# ¢7BçFFÆR/	ù:bçfVçF÷'"¢FbÒBäFFg&ÖRÆöEöçfVçF÷'¢bæ÷BFbæV×G ¢f3Âf3"Ò7Bæ6öÇVÖç2"¢6G2Ò²$ÆÂ%Ò²6÷'FVBFe²$6FVv÷'%ÒæG&÷æçVæVRçFöÆ7B¢6VÅö6BÒf3ç6VÆV7F&÷$6FVv÷'"Â6G2¢6V&6Òf3"çFWEöçWB%6V&6"Â"" ¢fÇFW&VBÒFbæ6÷¢b6VÅö6BÒ$ÆÂ# ¢fÇFW&VBÒfÇFW&VE¶fÇFW&VE²$6FVv÷'%ÒÓÒ6VÅö6EÐ¢b6V&6 ¢Ö6²Ò¢fÇFW&VE²%&öGV7BæÖR%Òç7G"æ6öçFç26V&6Â66SÔfÇ6RÂæÔfÇ6RÀ¢fÇFW&VE²%&öGV7B6öFR%Òç7G"æ6öçFç26V&6Â66SÔfÇ6RÂæÔfÇ6R¢¢fÇFW&VBÒfÇFW&VE¶Ö6µÐ ¢7BæFFg&ÖRfÇFW&VBÂW6Uö6öçFæW%÷vGFÕG'VRÂFUöæFWÕG'VR¢7Bæ6Föâb%6÷vær¶ÆVâfÇFW&VBÒöb¶ÆVâFbÒ&öGV7G2"¢VÇ6S ¢7Bçv&æær$æòçfVçF÷'FFâ"  ¢2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)HtS¢4ÄTåE2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H ¦VÆbvRÓÒ/	ùR6ÆVçG2# ¢7BçFFÆR/	ùR6ÆVçG2"¢FbÒBäFFg&ÖRÆöEö6ÆVçG2¢bæ÷BFbæV×G ¢6V&6Ò7BçFWEöçWB%6V&6æÖRÂ6ö×çÂ÷"6G"Â""¢fÇFW&VBÒFbæ6÷¢b6V&6 ¢6V&6&ÆRÒ²$6ÆVçBæÖR"Â$6ö×ç"Â$6G%Ð¢Ö6²ÒBå6W&W2fÇ6RÂæFWÖfÇFW&VBææFW¢f÷"6öÂâ6V&6&ÆS ¢b6öÂâfÇFW&VBæ6öÇVÖç3 ¢Ö6²ÒÖ6²ÂfÇFW&VE¶6öÅÒæ7GR7G"ç7G"æ6öçFç26V&6Â66SÔfÇ6RÂæÔfÇ6R¢fÇFW&VBÒfÇFW&VE¶Ö6µÐ¢7BæFFg&ÖRfÇFW&VBÂW6Uö6öçFæW%÷vGFÕG'VRÂFUöæFWÕG'VR¢7Bæ6Föâb%6÷vær¶ÆVâfÇFW&VBÒöb¶ÆVâFbÒ6ÆVçG2"¢VÇ6S ¢7Bçv&æær$æò6ÆVçBFFâ"  ¢2)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)HdôõDU")H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H)H §7Bç6FV&"æFfFW"§7Bç6FV&"æ6Föâb'cã(	B´$¥²væÖRu×Ò"§7Bç6FV&"æ6Föâb/	ù9â´$¥²wöæRu×Ò"
